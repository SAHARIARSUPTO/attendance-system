import os
import pickle
import numpy as np
import cv2
import face_recognition
import cvzone
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import shutil
import pygame
import time

# Initialize pygame mixer
pygame.mixer.init()

# Check if the encoding file exists
encode_file_path = 'local_data.pkl'
if not os.path.exists(encode_file_path):
    print(f"Error: Encoding file '{encode_file_path}' not found.")
    exit()

# Load the encoding file
with open(encode_file_path, 'rb') as file:
    encodeDict = pickle.load(file)

# Check if the background image exists
background_image_path = 'Resources/background.png'
if not os.path.exists(background_image_path):
    print(f"Error: Background image '{background_image_path}' not found.")
    exit()

# Initialize VideoCapture
cap = cv2.VideoCapture(0)
cap.set(3, 640)
cap.set(4, 480)

# Check if the camera is opened successfully
if not cap.isOpened():
    print("Error: Failed to open camera.")
    exit()

# Importing the mode images into a list
folderModePath = 'Resources/Modes'
if not os.path.exists(folderModePath):
    print(f"Error: Mode images folder '{folderModePath}' not found.")
    exit()

# Create a directory to store xls files
xls_directory = "Attendance_XLS"
if not os.path.exists(xls_directory):
    os.makedirs(xls_directory)

modePathList = os.listdir(folderModePath)
imgModeList = []
for path in modePathList:
    img_path = os.path.join(folderModePath, path)
    if os.path.isfile(img_path):
        imgModeList.append(cv2.imread(img_path))
    else:
        print(f"Warning: '{img_path}' is not a valid file.")

# Initialize variables
imgBackground = cv2.imread(background_image_path)
modeType = 0
counter = 0
id = -1
imgStudent = []
studentInfo = None

# Initialize attendance recording
class_name = input("Enter class name: ")
date = datetime.now().strftime("%Y-%m-%d")
attendance_file_name = os.path.join(xls_directory, f"{class_name}_{date}.xlsx")

# Initialize last attendance time
last_attendance_time = time.time()

# Function to reset attendance after 1 hour
def reset_attendance():
    global date, attendance_file_name, worksheet, workbook, start_time  # Declare start_time as global
    current_time = datetime.now()
    if current_time - start_time >= timedelta(hours=1):
        # Reset variables
        start_time = current_time
        date = current_time.strftime("%Y-%m-%d")
        attendance_file_name = os.path.join(xls_directory, f"{class_name}_{date}.xlsx")
        
        # Check if the attendance file already exists, if not, create it
        if not os.path.exists(attendance_file_name):
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["Date", "ID", "Attendance"])
            workbook.save(filename=attendance_file_name)
        else:
            workbook = load_workbook(filename=attendance_file_name)
            worksheet = workbook.active

# Check if the attendance file already exists, if not, create it
if not os.path.exists(attendance_file_name):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Date", "ID", "Attendance"])
    workbook.save(filename=attendance_file_name)
else:
    workbook = load_workbook(filename=attendance_file_name)
    worksheet = workbook.active

# Keep track of whether sounds have been played
attendance_sound_played = False
marked_sound_played = False
not_recognized_sound_played = False

# Function to play notification sound
def play_sound(sound_file):
    pygame.mixer.music.load(sound_file)
    pygame.mixer.music.play()

# Function to take attendance
def take_attendance(id, worksheet):
    global attendance_sound_played
    global marked_sound_played
    
    # Check if the student has already been marked present for this class
    for row in worksheet.iter_rows(values_only=True):
        if row[1] == id:
            if not marked_sound_played:
                play_sound(os.path.join("sounds", "already.wav"))
                marked_sound_played = True
            return False

    # If not, mark the attendance for this student
    worksheet.append([date, id, "Present"])
    workbook.save(filename=attendance_file_name)
    if not attendance_sound_played:
        play_sound(os.path.join("sounds", "success.wav"))
        attendance_sound_played = True
    return True

# Main loop
start_time = datetime.now()  # Initialize start time
while True:
    try:
        reset_attendance()  # Check if it's time to reset attendance
        
        # Read frame from the camera
        success, img = cap.read()

        if not success:
            print("Failed to read frame from the camera")
            break

        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        faceCurFrame = face_recognition.face_locations(imgS)
        encodeCurFrame = face_recognition.face_encodings(imgS, faceCurFrame)

        imgBackground[162:162 + 480, 55:55 + 640] = img

        match_found = False  # Flag to track if a match is found

        if faceCurFrame:
            for encodeFace, faceLoc in zip(encodeCurFrame, faceCurFrame):
                matches = face_recognition.compare_faces(list(encodeDict.values()), encodeFace)
                faceDis = face_recognition.face_distance(list(encodeDict.values()), encodeFace)

                matchIndex = np.argmin(faceDis)

                if 0 <= matchIndex < len(matches) and matches[matchIndex]:
                    match_found = True  # Set the flag to True
                    id = list(encodeDict.keys())[matchIndex]

                    # Take attendance for the detected student
                    current_time = time.time()
                    if current_time - last_attendance_time >= 5:  # 5 seconds delay
                        success = take_attendance(id, worksheet)
                        last_attendance_time = current_time

                        # Load image corresponding to the detected ID
                        img_path = f'images/{id}.png'  # Assuming images are named after IDs
                        if os.path.exists(img_path):
                            img_student = cv2.imread(img_path)
                            img_student = cv2.resize(img_student, (414, 633))  # Resize to match the display area
                            imgBackground[44:44 + 633, 808:808 + 414] = img_student

                        studentInfo = encodeDict.get(id)
                        if studentInfo is not None:
                            if success:
                                cv2.putText(imgBackground, "Attendance Recorded", (50, 50),
                                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2, cv2.LINE_AA)

                else:
                    # Play "not recognized" sound
                    if not not_recognized_sound_played:
                        play_sound(os.path.join("sounds", "not.wav"))
                        not_recognized_sound_played = True

                y1, x2, y2, x1 = faceLoc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                bbox = 55 + x1, 162 + y1, x2 - x1, y2 - x1
                imgBackground = cvzone.cornerRect(imgBackground, bbox, rt=0)

                if counter == 0:
                    cvzone.putTextRect(imgBackground, "Loading", (275, 400))
                    cv2.imshow("Face Attendance", imgBackground)
                    cv2.waitKey(1)
                    counter = 1
                    modeType = 1

            if counter != 0 and match_found:
                if counter == 1:
                    pass
                pass

        else:
            modeType = 0
            counter = 0

        # Reset sound flags if no face is detected
        if not faceCurFrame:
            attendance_sound_played = False
            marked_sound_played = False
            not_recognized_sound_played = False

        cv2.imshow("Face Attendance", imgBackground)
        cv2.waitKey(1)

    except Exception as e:
        print("Error:", e)
        break

cap.release()
cv2.destroyAllWindows()
